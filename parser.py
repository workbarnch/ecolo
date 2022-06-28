from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import json
import os
import time

def generate_excel(data):
    coeff = data.get('coeff')
    file_name = f"{data.get('questionnaire')}.xlsx"
    file_exists = os.path.exists(f"excel/{file_name}")
    if file_exists:
        try:
            wb = load_workbook(f"excel/{file_name}", data_only=False)
            wb.active = 0
            sheet = wb.active
            for idx, row in enumerate(sheet.rows, start=1):
                row[3].value = coeff

            wb.active = 6
            sheet = wb.active
            for index, row in enumerate(sheet.rows, start=1):
                for q in data.get('q'):
                    if row[2].value == q:
                        sheet.cell(row=index, column=4).value = data.get('q').get(q)
            wb.active = 6
            sheet = wb.active
            sheet['N7'].value = 1
            wb.save(file_name)
            time.sleep(1)
            os.system('rm -rf tmp')
            os.system(f"libreoffice --headless --convert-to xlsx {file_name} --outdir tmp")
            time.sleep(1)
            return read_excel_data(file_name)
        except Exception as _ex:
            print(_ex);
    else:
        return {
            "status": False
        }


def read_excel_data(file_name):
    file_exists = os.path.exists(f"tmp/{file_name}")

    if file_exists:
        wb = load_workbook(f"tmp/{file_name}", data_only=True)
        wb.active = 9
        sheet = wb.active
        data = {
            "status": True,
            "normality": sheet['B20'].value,
            "coherence": sheet['B21'].value,
            "estimated_revenue_lower": sheet['B33'].value,
            "estimated_revenue_upper": sheet['C33'].value,
            "estimated_revenue_client": sheet['D33'].value,
            "estimated_profit_lower": sheet['B34'].value,
            "estimated_profit_upper": sheet['C34'].value,
            "estimated_profit_client": sheet['D34'].value,
            "deviation_degree": sheet['B24'].value,
        }
        os.system(f'rm -rf tmp {file_name} ')
        return data
    else:
        return {
            'status': False
        }


def main(data):
    return generate_excel(data)
